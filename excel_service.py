import openpyxl


def extract_data(path):
    wb = openpyxl.load_workbook(path)
    sheetname = wb.get_sheet_names()[0]
    sheet = wb.get_sheet_by_name(sheetname)
    new_sheet = wb.create_sheet('Extracted Data')
    # copy column a
    copy_column(0, sheet, new_sheet)
    wb.save(path)


def copy_column(column_index, from_sheet, to_sheet):
    col_data = list(from_sheet.columns)[column_index]  # 0-indexing

    for idx, cell in enumerate(col_data, 1):
        to_sheet.cell(row=idx,
                      column=column_index + 1).value = cell.value  # 1-indexing
